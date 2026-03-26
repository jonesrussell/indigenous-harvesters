"""
Export community enrichment coverage to Excel spreadsheet.

Run: .venv/bin/python notebooks/export_coverage_spreadsheet.py

Requires SSH tunnel: ssh -f -N -L 18050:localhost:8050 deployer@147.182.150.145
"""

import json
from datetime import datetime, timezone
from pathlib import Path

import httpx
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DATA_DIR = Path(__file__).parent / "data"
BASE = "https://api.northcloud.one"


def fetch_all_communities():
    all_nc = []
    offset = 0
    with httpx.Client(timeout=60) as client:
        while True:
            resp = client.get(f"{BASE}/api/v1/communities?limit=100&offset={offset}")
            resp.raise_for_status()
            batch = resp.json()["communities"]
            all_nc.extend(batch)
            if len(batch) < 100:
                break
            offset += 100
    return all_nc


def fetch_all_band_offices(communities):
    offices = {}
    with httpx.Client(timeout=30) as client:
        for i, c in enumerate(communities):
            cid = c["id"]
            try:
                resp = client.get(f"{BASE}/api/v1/communities/{cid}/band-office")
                if resp.status_code == 200:
                    offices[cid] = resp.json().get("band_office") or {}
                else:
                    offices[cid] = {}
            except Exception:
                offices[cid] = {}

            if (i + 1) % 100 == 0:
                print(f"  Band offices: {i + 1}/{len(communities)}...")

    return offices


def fetch_people_count(communities):
    counts = {}
    with httpx.Client(timeout=30) as client:
        for i, c in enumerate(communities):
            cid = c["id"]
            try:
                resp = client.get(f"{BASE}/api/v1/communities/{cid}/people?limit=1&offset=0")
                if resp.status_code == 200:
                    data = resp.json()
                    counts[cid] = data.get("total", 0)
                else:
                    counts[cid] = 0
            except Exception:
                counts[cid] = 0

            if (i + 1) % 100 == 0:
                print(f"  People: {i + 1}/{len(communities)}...")

    return counts


def main():
    print("Fetching communities...")
    communities = fetch_all_communities()
    communities.sort(key=lambda c: (c.get("province", ""), c.get("name", "")))
    print(f"Total: {len(communities)}")

    print("Fetching band offices...")
    offices = fetch_all_band_offices(communities)

    print("Fetching people counts...")
    people_counts = fetch_people_count(communities)

    # Create workbook
    wb = openpyxl.Workbook()

    # --- Summary sheet ---
    ws_summary = wb.active
    ws_summary.title = "Summary"

    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    green_fill = PatternFill(start_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    ws_summary.append(["Community Data Enrichment — Coverage Report"])
    ws_summary["A1"].font = Font(bold=True, size=14)
    ws_summary.append([f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"])
    ws_summary.append([])

    # Province summary
    summary_headers = ["Province", "Total", "Email", "Email %", "Website", "Phone", "Chief/Council", "Missing Email"]
    ws_summary.append(summary_headers)
    for col_idx, _ in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=4, column=col_idx)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border

    by_province = {}
    for c in communities:
        prov = c.get("province", "??")
        if prov not in by_province:
            by_province[prov] = {"total": 0, "email": 0, "website": 0, "phone": 0, "people": 0}
        by_province[prov]["total"] += 1
        office = offices.get(c["id"], {})
        if office.get("email"):
            by_province[prov]["email"] += 1
        if c.get("website"):
            by_province[prov]["website"] += 1
        if office.get("phone"):
            by_province[prov]["phone"] += 1
        if people_counts.get(c["id"], 0) > 0:
            by_province[prov]["people"] += 1

    row = 5
    for prov in sorted(by_province.keys()):
        d = by_province[prov]
        pct = d["email"] / d["total"] if d["total"] else 0
        ws_summary.append([
            prov, d["total"], d["email"], f"{pct:.0%}",
            d["website"], d["phone"], d["people"],
            d["total"] - d["email"],
        ])
        # Color code email %
        pct_cell = ws_summary.cell(row=row, column=4)
        if pct >= 0.6:
            pct_cell.fill = green_fill
        elif pct >= 0.4:
            pct_cell.fill = yellow_fill
        else:
            pct_cell.fill = red_fill
        for col_idx in range(1, len(summary_headers) + 1):
            ws_summary.cell(row=row, column=col_idx).border = thin_border
        row += 1

    # Totals row
    totals = {"total": 0, "email": 0, "website": 0, "phone": 0, "people": 0}
    for d in by_province.values():
        for k in totals:
            totals[k] += d[k]
    pct = totals["email"] / totals["total"] if totals["total"] else 0
    ws_summary.append([
        "TOTAL", totals["total"], totals["email"], f"{pct:.0%}",
        totals["website"], totals["phone"], totals["people"],
        totals["total"] - totals["email"],
    ])
    for col_idx in range(1, len(summary_headers) + 1):
        cell = ws_summary.cell(row=row, column=col_idx)
        cell.font = Font(bold=True)
        cell.border = thin_border

    # Auto-width
    for col_idx in range(1, len(summary_headers) + 1):
        ws_summary.column_dimensions[get_column_letter(col_idx)].width = 15

    # --- Detail sheet ---
    ws_detail = wb.create_sheet("All Communities")

    detail_headers = [
        "Name", "Province", "Population", "Website", "Email",
        "Phone", "Fax", "Address", "City", "Postal Code",
        "Chief/Council Count", "Data Source", "INAC ID",
    ]
    ws_detail.append(detail_headers)
    for col_idx, _ in enumerate(detail_headers, 1):
        cell = ws_detail.cell(row=1, column=col_idx)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border

    for c in communities:
        office = offices.get(c["id"], {})
        pcount = people_counts.get(c["id"], 0)
        ws_detail.append([
            c.get("name", ""),
            c.get("province", ""),
            c.get("population", ""),
            c.get("website", ""),
            office.get("email", ""),
            office.get("phone", ""),
            office.get("fax", ""),
            office.get("address_line1", ""),
            office.get("city", ""),
            office.get("postal_code", ""),
            pcount if pcount > 0 else "",
            office.get("data_source", ""),
            c.get("inac_id", ""),
        ])

    # Auto-width detail
    col_widths = [35, 8, 12, 40, 35, 18, 18, 35, 20, 12, 16, 15, 10]
    for i, w in enumerate(col_widths, 1):
        ws_detail.column_dimensions[get_column_letter(i)].width = w

    # Auto-filter
    ws_detail.auto_filter.ref = f"A1:{get_column_letter(len(detail_headers))}{len(communities) + 1}"

    # --- Missing Email sheet ---
    ws_missing = wb.create_sheet("Missing Email")

    missing_headers = ["Name", "Province", "Has Website", "Website URL", "Has Phone", "Phone"]
    ws_missing.append(missing_headers)
    for col_idx, _ in enumerate(missing_headers, 1):
        cell = ws_missing.cell(row=1, column=col_idx)
        cell.font = header_font_white
        cell.fill = PatternFill(start_color="C0504D", end_color="C0504D", fill_type="solid")
        cell.border = thin_border

    for c in communities:
        office = offices.get(c["id"], {})
        if not office.get("email"):
            ws_missing.append([
                c.get("name", ""),
                c.get("province", ""),
                "Yes" if c.get("website") else "No",
                c.get("website", ""),
                "Yes" if office.get("phone") else "No",
                office.get("phone", ""),
            ])

    missing_widths = [35, 8, 12, 40, 12, 18]
    for i, w in enumerate(missing_widths, 1):
        ws_missing.column_dimensions[get_column_letter(i)].width = w
    ws_missing.auto_filter.ref = f"A1:{get_column_letter(len(missing_headers))}{ws_missing.max_row}"

    # Save
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out = DATA_DIR / f"community_enrichment_review_{timestamp}.xlsx"
    wb.save(out)
    print(f"\nSaved: {out}")

    # Also save as the canonical name
    canonical = DATA_DIR / "community_enrichment_review.xlsx"
    wb.save(canonical)
    print(f"Also saved: {canonical}")


if __name__ == "__main__":
    main()
